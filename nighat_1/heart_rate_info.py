import openpyxl
from random import shuffle
import random
import numpy as np
from openpyxl import load_workbook

# set random seed
random.seed(0)

# CONSTANTS
NUM_DATA = 130
EPOCHS = 1
TRAIN_NUM = 100


class perceptron:

    def __init__(self, num_data, workbook_name, percentage_split):
        # load the workbook to read the data from
        self.wb = load_workbook(workbook_name)
        self.ws = self.wb.active

        # the input data will be heart-rate, and sex
        self.input_data = [None] * num_data
        self.correct_values = [None] * num_data
        self.data = [None] * num_data

        # save the num_data variable
        self.num_data = num_data
        self.percentage_split = percentage_split

        # create the weights and bias
        self.weights = np.array([0.0] *2) # sex and heart-rate
        self.bias = 0
        self.accuracy = 0

        # calculate the training split
        self.train_num = int(np.floor(percentage_split * self.num_data))

    def normalize_vector(self, vector):
        mag = self.get_manitude(vector)

        return vector / mag

    def get_manitude(self, vector):
        squared = vector ** 2

        sum = 0
        for element in squared:
            sum += element

        mag = sum ** (1 / 2)

        return mag

    def read_data(self, random_seed):
        # read data
        self.random_seed = random_seed
        for line_num in range(1, self.num_data + 1):
            list_index = line_num - 1
            heart_rate = self.ws['D' + str(line_num)].value
            if float(self.ws['B' + str(line_num)].value) >= 100.00:
                fever = 1
            else:
                fever = -1
            sex = self.ws['C' + str(line_num)].value
            input_data  = np.array([float(sex), float(heart_rate)])
            correct_value = float(fever)
            self.data[list_index] = [input_data, correct_value]

        random.seed(random_seed)
        shuffle(self.data)

        # normalize the data
        for i in range(self.num_data):
            self.data[i][0] = self.normalize_vector(self.data[i][0])

    def train(self, epochs):
        # train
        for e in range(epochs):
            for sample_index in range(TRAIN_NUM):
                if np.dot(self.weights, self.data[sample_index][0]) * self.data[sample_index][1] <= 0:
                    self.weights += self.data[sample_index][1] * self.data[sample_index][0]
                    self.bias += self.data[sample_index][1]


    def test(self):
        # track the number of sample types
        self.fevers = 0
        self.not_fevers = 0

        self.num_correct = 0
        self.num_total = 0
        for test_index in range(self.train_num, self.num_data):
            prediction = np.dot(self.data[test_index][0], self.weights) + self.bias
            if prediction > 0:
                prediction = 1
            else:
                prediction = -1

            self.num_total += 1
            if prediction == self.data[test_index][1]:
                self.num_correct += 1

            if self.data[test_index][1] == -1:
                self.not_fevers += 1
            else:
               self.fevers += 1

        if not (self.fevers == 0 or self.not_fevers == 0):
            self.accuracy = float(self.num_correct) / float(self.num_total)

    def print(self):
        if not(self.fevers == 0 or self.not_fevers == 0):
            print("##### random seed: ", self.random_seed, "  split: ", self.percentage_split)
            print("Number of fevers in test sample: ",self.fevers)
            print("Number of not fevers in test sample: ", self.not_fevers)
            print("Num correct: ", self.num_correct, '\nNum Total: ', self.num_total)
            print(float(self.num_correct) / float(self.num_total), '\n\n')



num_random_seeds = 10
split_start = 60
split_end = 96
jump = 5
num_splits = int(np.floor((split_end - split_start) / jump)) + 1

perceptrons = [None] * (num_splits * num_random_seeds)
current_index = 0
for i in range(num_random_seeds):
    for j in range(split_start, split_end, jump):
        percep = perceptron(130, 'pickles/hr.xlsx', j * .01)
        percep.read_data(i)
        percep.train(j)
        percep.test()
        perceptrons[current_index] = percep
        current_index += 1

perceptrons = sorted(perceptrons, key=lambda x: x.accuracy, reverse=True)
print(perceptrons)
for i in range(10):
    perceptrons[i].print()
