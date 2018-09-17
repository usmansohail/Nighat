import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import pickle
from classes import symbol, character, BKNode, levenshtein
from difflib import SequenceMatcher
from openpyxl import Workbook
from get_symbol_and_characters import CHARS_DICT_NAME, CHARS_LIST_NAME, SYMBOL_DICT_NAME, SYMBOLS_LIST_NAME

# constants

SYMS_PICKLE_NAME = "pickles/syms-3.p"
COMPOSITION_WB_NAME = "pickles/word_composition.xlsx"

# This file will implement methods for getting the symbols from either the ID, or
# the composition words


# import the characters and symbols
chars = pickle.load(open(CHARS_LIST_NAME, 'rb'))
# chars_dict = pickle.load(open(CHARS_DICT_NAME, 'rb'))


syms = pickle.load(open(SYMBOLS_LIST_NAME, 'rb'))
syms_dict = pickle.load(open(SYMBOL_DICT_NAME, 'rb'))



# words = []
#
# for sym in syms:
#     words.append(sym.words[0])
#
# bk_dict = BKNode("")
#
# for word in words:
#     bk_dict.insert(word)
#
#
# results = []
# bk_dict.search("run", 1, results)
#
# print(results)

bk_id_dict = BKNode("", None)
def get_bk_dict():
    for sym in syms:
        for word in sym.words:
            bk_id_dict.insert(word, sym.id)

    #results = []

    #bk_id_dict.search("run", 1, results)

    #print(results)


wb = load_workbook('pickles/dictionary.xlsx')



def match_longest_string(string1, string2):
    match = SequenceMatcher(None, string1, string2).find_longest_match(0, len(string1), 0, len(string2))
    return match[2]


match_longest_string("run", "to run")
match_longest_string("running", "running to")
match_longest_string("run", "runger")
match_longest_string("run", "brung")


def get_most_likely_symbols(phrase):
    # first get results from BK tree
    results = []
    bk_id_dict.search(phrase, 5, results)

    # of these results, order them by longest common substring
    temp = [None]*len(results)

    for i in range(len(results)):
        # print('w1: ', phrase, ' w2: ', results[i][1])
        # compute score, longest_common_substring / distance
            # add one for the distance case of 0
        score = float(match_longest_string(phrase, results[i][1])) / float(results[i][0] + 1)
        temp[i] = (score, results[i])

    sorted_results = sorted(temp, key=lambda x: x[0], reverse=True)

    # print(phrase, ": \n")
    # for i in range(5):
    #     if i < len(sorted_results):
    #         print("Score: ", sorted_results[i][0], "word: ", sorted_results[i][1])
    # print('\n\n')

    if len(sorted_results) > 0:
        final_result = sorted_results[0]
    else:
        final_result = None

    return final_result

get_bk_dict()
# (get_most_likely_symbols('gun'))
# (get_most_likely_symbols('run'))
# (get_most_likely_symbols('to expect'))
# (get_most_likely_symbols('anticipate'))

#get_most_likely_symbols("description indicator")


def add_id():

    num_total = len(syms)
    curr_num = 1

    # display all composition
    for sym in syms:
        # print(sym.words, ": ", sym.composition)
        likely_id = []
        sym_id = []
        # if len(sym.composition) > 1:
        #     for word in sym.composition:
        #         if word is not '':
        #             likely_id = get_most_likely_symbols(word)
        #             if len(likely_id) > 0:
        #                 sym_id.append(likely_id[1][2])
        #                 print("adding: ", likely_id[1][1], " to ", sym.words)
        # if len(sym_id) > 0:
        #     sym.id_composition = sym_id

        not_a_char = not sym.is_character
        current_symbol = sym

        print("################### word: ", sym.words, '       num: ', curr_num, '/', num_total)
        curr_num += 1

        while not_a_char:
            print("the symbol is not a character, searching for composition")
            print("words: ", sym.words)
            print("id: ", sym.id)
            print("word comp: ", current_symbol.composition, '\n\n')
            sym_id.append('[')
            for word in current_symbol.composition:
                if word is not '':
                    likely_id = get_most_likely_symbols(word)
                    if likely_id is not None:
                        temp_symbol = syms_dict[likely_id[1][2]]
                        sym_id.append(temp_symbol.id)
                        print("adding: ", likely_id[1][1], ' to ', sym.words)
                        print("adding: ", likely_id[1][2], ' to ', sym_id)

                        if temp_symbol.is_character or len(temp_symbol.composition) == 0:
                            not_a_char = False

                        current_symbol = syms_dict[temp_symbol.id]
                    else:
                        not_a_char = False
                else:
                    not_a_char = False

            sym_id.append(']')

        if len(sym_id) > 0:
            sym.id_composition = sym_id

        print('\n\n\n\n')

        #
        # print(sym.words)
        # print(sym.composition)
        # print(sym.id_composition, '\n')

    pickle.dump(syms, open(SYMS_PICKLE_NAME, 'wb'))


add_id()




syms = pickle.load(open(SYMS_PICKLE_NAME, 'rb'))

def make_wb(book_name):
    # save the id's to a workbook
    wb = Workbook()
    sheet = wb.active

    pickle.dump(syms, open(SYMS_PICKLE_NAME, 'wb'))

    num_not_complete = 0
    # get id_to_words_dict
    id_to_words = pickle.load(open("id_to_words.p", 'rb'))

    for i in range(0, len(syms)):
        sheet['A' + str(i + 1)] = syms[i].id
        sheet['B' + str(i + 1)] = ", ".join(list(syms[i].words))
        sheet['C' + str(i + 1)] = ", ".join(list(syms[i].composition))
        sheet['D' + str(i + 1)] = ", ".join(str(sym_id) for sym_id in list(syms[i].id_composition))

        temp = []
        for id in syms[i].id_composition:
            temp.append(id_to_words[id])
            temp.append(';\n')

        sheet['E' + str(i + 1)] = ", ".join(temp)

        if len(list(syms[i].composition)) is not len(list(syms[i].id_composition)) and \
            len(list(syms[i].composition)) > 1:
            sheet['F' + str(i + 1)] = "***"
            num_not_complete += 1

    wb.save(book_name)
    print("Number of incomplete symbols: ", num_not_complete)

make_wb(COMPOSITION_WB_NAME)