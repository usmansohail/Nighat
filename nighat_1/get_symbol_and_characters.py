import openpyxl as pyxl
from openpyxl import worksheet
import pickle
import re
from classes import symbol, character


# constants
CHARS_LIST_NAME = "pickles/chars.p"
CHARS_DICT_NAME = "pickles/chars_dict.p"

SYMBOLS_LIST_NAME = "pickles/symbols.p"
SYMBOL_DICT_NAME = 'pickles/symbols_dict.p'

# get all the characters
def get_characters():
    # open the excel file
    file = open("pickles/dictionary-n.xlsx", 'rb')
    wb = pyxl.load_workbook(file)
    ws = wb.active

    # store characters and words
    words = []
    characters = []
    characters_dict = {}

    symbols = []
    symbols_dict = {}



    lines = []
    for row in ws.iter_rows():
        if "BCI-AV#" not in str(row[0].value) and row[0].value is not None:
            lines.append((row[0].value, str(row[2].value), str(row[3].value)))
            #print(lines[-1])

    # for row in ws.iter_rows():
    #     if "BCI-AV#" not in str(row[0].value) and row[0].value is not None:
    #         if str(row[2].value) is not "":
    #             lines.append((row[0].value, str(row[1].value), str(row[2].value)))
    #         else:
    #             lines.append((row[0].value, str(row[1].value), str(row[3].value)))
            #print(lines[-1])

    # create dicts
    num_to_def_dict = {}
    for line in lines:
        words_non_filtered = str(line[1]).split(',')
        words_filtered = []
        # print(words_non_filtered)
        for word in words_non_filtered:
            regex = re.compile(".*?\((.*?)\)")
            remove_string = re.findall(regex, word)
            if len(remove_string) > 0:
                word = word.replace(remove_string[0], '')
                word = word.replace('(', '')
                word = word.replace(')', '')
            if len(word) > 0 and word[-1] is '_':
                word_list = list(word)
                word_list[-1] = ''
                word = "".join(word_list[:-1])
            word = word.replace('_', ' ')
            words_filtered.append(word)
        # print(words_filtered)
        num_to_def_dict[line[0]] = ()

    # "BCI-AV#"
    # for i in range(1000, 1005):
    #     print(lines[i])
    #
    # print(lines[-1])

    id_to_words_dict = {}
    for line in lines:
        id_to_words_dict[line[0]] = line[1]




    def_dict = {}
    id_to_character_dict = {}
    for line in lines:
        def_dict[line[0]] = line[2]

        composition = ""

        # assume it is a word, unless it expresses that it is a character
        is_character = False

        # get the description/composition
        if " - Character" in str(line[2]):
            id_to_character_dict[line[0]] = [line[1], line[2]]
            is_character = True

        if "+" in str(line[2]):
            regex = re.compile(r"\(.+\)")
            composition = str(re.findall(regex, str(line[2]).replace('\n', '')))
            composition = composition.replace('(', '')
            composition = composition.replace(')', '')
            composition = composition.replace('\'', '')
            composition = composition.replace('"', '')
            # composition = composition.replace('_', ' ')



            # remove [ first bracket
            composition = composition[1:-1]




            #print("Comp: ", composition)
            regex_2 = re.compile(r":.*")
            remove_string = re.findall(regex_2, str(composition))
            while len(remove_string) > 0:
                for i in range(len(remove_string)):
                    composition = composition.replace(remove_string[0], '')
                remove_string = re.findall(regex_2, composition)
                # print("parts removed: ", remove_string[0])
            regex_2 = re.compile(r": .*")
            remove_string = re.findall(regex_2, str(composition))
            while len(remove_string) > 0:
                for i in range(len(remove_string)):
                    composition = composition.replace(remove_string[i], '')
                remove_string = re.findall(regex_2, composition)
                #print("parts removed: ", remove_string[0])
            while ',' in composition:
                regex = re.compile(r"(,\w*_*\w*)*")
                remove_string = re.findall(regex, composition)
                for i in range(len(remove_string)):
                    composition = composition.replace(remove_string[i], '')
                    #print("remove string:", remove_string)
            regex = re.compile(r"\[[^\+]*\]")
            remove_string = re.findall(regex, composition)
            reg2 = re.compile(r"\-.*[cC]haracter.*")
            remove = re.findall(reg2, composition)
            while len(remove_string) > 0:
                for i in range(len(remove_string)):
                    composition = composition.replace(remove_string[i], '')
                remove_string = re.findall(regex, composition)
            while len(remove) > 0:
                for i in range(len(remove)):
                    composition = composition.replace(remove[i], '')
                remove = re.findall(reg2, composition)
            print("Comp: ", composition)
            # print("final string: ", composition)
            # print("part 1: ", line[0], '\n', "part 2: ", line[1], '\n', "part 3: ", line[2], '\n')
            # if composition == "":
                # print('wtf', line[0], line[1], line[2])

        # print(line[0], ": ", composition, " ; ", line)


        if(is_character):
            t_is_word = not is_character
            t_char = symbol(line[1], composition.split(" + "), line[0], t_is_word)
            characters.append(t_char)
            characters_dict[t_char.id] = t_char



        sym = symbol(line[1], composition.split(" + "), line[0], not is_character)
        symbols.append(sym)
        symbols_dict[sym.id] = sym


    #print(len(id_to_character_dict.keys()))

    pickle.dump(id_to_character_dict, open("bliss_chars.p", 'wb'))


    word_to_id_dict = {}
    ambiguous_words = {}
    for key in id_to_words_dict.keys():
        words = str(id_to_words_dict[key]).split(',')
        for word in words:
            if word in word_to_id_dict.keys():
                # add it to the ambigous words dictionary
                if word in ambiguous_words.keys():
                    ambiguous_words[word].append(key)
                else:
                    ambiguous_words[word] = [key]
                    ambiguous_words[word].append(word_to_id_dict[word][0])
                word_to_id_dict[word].append(key)
                #print(word, "  ", word_to_id_dict[word])
            else:
                word_to_id_dict[word] = [key]

    #print(len(word_to_id_dict.keys()))
    pickle.dump(word_to_id_dict, open("word_to_char_id.p", "wb"))

    pickle.dump(characters, open(CHARS_LIST_NAME, 'wb'))
    pickle.dump(symbols, open(SYMBOLS_LIST_NAME, 'wb'))
    pickle.dump(symbols_dict, open(SYMBOL_DICT_NAME, 'wb'))

    pickle.dump(id_to_words_dict, open("id_to_words.p", 'wb'))







get_characters()
chars  = pickle.load(open(CHARS_LIST_NAME, 'rb'))
syms  = pickle.load(open(CHARS_LIST_NAME, 'rb'))

# num_indicators = 0
# num_char_ind = 0
#
def ind_info():
    for sym in syms:
        if not sym.ind():
            sym.display_info()
#         num_indicators += 1
# for sym in chars:
#     if sym.ind():
#         sym.display_info
#         num_indicators += 1
#
# print("num indicators: ", num_indicators)
# print("chars: ", len(chars))
# print("syms: ", len(syms))
# print("syms and chars: ", len(syms) + len(chars))
#


def get_dicts():
    # create a dictionary from id to symbol, as well as a dictionary from composition ids to symbol
    id_to_symbol = {}
    composition_to_symbol = {}
    def_to_symbol = {}
    for sym in syms:
        id_to_symbol[sym.id] = sym
        # def_to_symbol[sym.words] =
        # if len(sym.composition) > 0:
